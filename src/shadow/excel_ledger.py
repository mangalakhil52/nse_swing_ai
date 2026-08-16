"""
Excel Trade Ledger Manager Module.
Maintains a professional, auto-updating Excel spreadsheet (data/trade_ledger.xlsx) recording all trade recommendations,
daily EOD price updates, target/stop-loss status tracking, holding period count, and realized/unrealized P&L.
"""

from datetime import date, datetime
import logging
from pathlib import Path
from typing import Any
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.settings import settings
from src.core.models import TradeRecommendation
from src.core.types import ExitReason, TradeStatus

logger = logging.getLogger(__name__)


class TradeLedgerExcelManager:
    """Manages auto-updating Excel trade tracking ledger."""

    COLUMNS = [
        "Recommendation ID",
        "Date",
        "Symbol",
        "Company Name",
        "Sector",
        "Conviction",
        "Score",
        "CMP (₹)",
        "Entry Trigger (₹)",
        "Stop Loss (₹)",
        "Target 1 (₹)",
        "Target 2 (₹)",
        "Target 3 (₹)",
        "Shares",
        "Capital Allocated (₹)",
        "Current Price (₹)",
        "Status",
        "Holding Sessions",
        "P&L (%)",
        "P&L (₹)",
        "Exit Date",
        "Exit Price (₹)",
        "Exit Reason",
        "Last Updated",
    ]

    def __init__(self, file_path: Path | None = None):
        self.file_path = file_path or settings.DATA_DIR / "trade_ledger.xlsx"
        self.csv_path = settings.DATA_DIR / "trade_ledger.csv"
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        self._ensure_file_exists()

    def _ensure_file_exists(self) -> None:
        """Initializes empty ledger Excel file if not already present."""
        if not self.file_path.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trade Ledger"
            ws.append(self.COLUMNS)
            self._apply_header_styles(ws)
            wb.save(self.file_path)

            # CSV fallback
            df = pd.DataFrame(columns=self.COLUMNS)
            df.to_csv(self.csv_path, index=False)
            logger.info(f"Initialized new Excel trade ledger at {self.file_path}")

    def record_recommendations(self, recs: list[TradeRecommendation]) -> int:
        """Appends new actionable recommendations to the ledger."""
        if not recs:
            return 0

        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Trade Ledger"]

        # Get existing recommendation IDs to avoid duplicate rows
        existing_ids = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                existing_ids.add(str(row[0]))

        added_count = 0
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for rec in recs:
            if rec.recommendation_id in existing_ids:
                continue

            row_data = [
                rec.recommendation_id,
                str(rec.recommendation_date),
                rec.symbol,
                rec.company_name,
                rec.sector,
                rec.conviction.value,
                round(rec.composite_score, 1),
                round(rec.levels.current_market_price, 2),
                round(rec.levels.entry_trigger_price, 2),
                round(rec.levels.stop_loss_price, 2),
                round(rec.levels.target_1, 2),
                round(rec.levels.target_2, 2),
                round(rec.levels.target_3, 2),
                rec.levels.position_size_shares,
                round(rec.levels.allocated_capital_rupees, 2),
                round(rec.levels.current_market_price, 2),  # Current Price initially = CMP
                TradeStatus.PENDING_ENTRY.value,
                0,     # Holding Sessions
                0.0,   # P&L %
                0.0,   # P&L ₹
                "",    # Exit Date
                "",    # Exit Price
                "",    # Exit Reason
                now_str,
            ]
            ws.append(row_data)
            added_count += 1

        if added_count > 0:
            self._apply_sheet_formatting(ws)
            wb.save(self.file_path)
            self._sync_to_csv()
            logger.info(f"Recorded {added_count} new trade(s) in Excel ledger.")

        return added_count

    def update_ledger_with_eod_data(self, bhavcopy_df: pd.DataFrame, scan_date: date) -> str:
        """
        Evaluates EOD price action from Bhavcopy against all open trades in the Excel ledger.
        Updates status, unrealized/realized P&L, holding sessions, and exit metrics.
        """
        if not self.file_path.exists() or bhavcopy_df.empty:
            return "No open trades or empty Bhavcopy data."

        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Trade Ledger"]

        # Map symbol -> row dictionary for fast lookup
        bhav_lookup = {}
        for _, row in bhavcopy_df.iterrows():
            sym = str(row.get("symbol", "")).upper().strip()
            if sym:
                bhav_lookup[sym] = {
                    "open": float(row.get("open", row.get("close", 0.0))),
                    "high": float(row.get("high", row.get("close", 0.0))),
                    "low": float(row.get("low", row.get("close", 0.0))),
                    "close": float(row.get("close", 0.0)),
                }

        updated_summary = []
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for row_idx in range(2, ws.max_row + 1):
            status = str(ws.cell(row=row_idx, column=17).value or "")

            # Only process non-terminal trades
            if status in [
                TradeStatus.STOPPED_OUT.value,
                TradeStatus.TARGET_3_HIT.value,
                TradeStatus.TIME_EXPIRED.value,
                TradeStatus.CANCELLED.value,
            ]:
                continue

            symbol = str(ws.cell(row=row_idx, column=3).value or "").strip().upper()
            bar = bhav_lookup.get(symbol)
            if not bar or bar["close"] <= 0:
                continue

            entry_trigger = float(ws.cell(row=row_idx, column=9).value or 0.0)
            stop_loss = float(ws.cell(row=row_idx, column=10).value or 0.0)
            target_1 = float(ws.cell(row=row_idx, column=11).value or 0.0)
            target_2 = float(ws.cell(row=row_idx, column=12).value or 0.0)
            target_3 = float(ws.cell(row=row_idx, column=13).value or 0.0)
            shares = int(ws.cell(row=row_idx, column=14).value or 1)
            holding_sessions = int(ws.cell(row=row_idx, column=18).value or 0) + 1

            close_p = bar["close"]
            high_p = bar["high"]
            low_p = bar["low"]
            open_p = bar["open"]

            # Update Current Price cell
            ws.cell(row=row_idx, column=16, value=round(close_p, 2))
            ws.cell(row=row_idx, column=18, value=holding_sessions)
            ws.cell(row=row_idx, column=24, value=now_str)

            # 1. Transition PENDING_ENTRY -> ACTIVE if high >= entry_trigger
            if status == TradeStatus.PENDING_ENTRY.value:
                if high_p >= entry_trigger * 0.998:
                    status = TradeStatus.ACTIVE.value
                    ws.cell(row=row_idx, column=17, value=status)
                    logger.info(f"[LEDGER] {symbol} triggered entry at ₹{entry_trigger:.2f} -> ACTIVE")
                    updated_summary.append(f"🟢 *{symbol}*: Triggered Entry at ₹{entry_trigger:.2f} (Status: ACTIVE)")
                else:
                    # Still pending, update price and continue
                    continue

            # 2. Check Stop Loss Execution
            if open_p < stop_loss:
                # Gapped down through stop
                status = TradeStatus.STOPPED_OUT.value
                exit_price = open_p
                exit_reason = ExitReason.STOP_LOSS.value
                pnl_rupees = (exit_price - entry_trigger) * shares
                pnl_pct = ((exit_price - entry_trigger) / entry_trigger) * 100.0

                ws.cell(row=row_idx, column=17, value=status)
                ws.cell(row=row_idx, column=19, value=round(pnl_pct, 2))
                ws.cell(row=row_idx, column=20, value=round(pnl_rupees, 2))
                ws.cell(row=row_idx, column=21, value=str(scan_date))
                ws.cell(row=row_idx, column=22, value=round(exit_price, 2))
                ws.cell(row=row_idx, column=23, value=exit_reason)
                updated_summary.append(f"🔴 *{symbol}*: Gapped through Stop Loss at ₹{exit_price:.2f} (P&L: {pnl_pct:+.2f}%)")
                continue

            if low_p <= stop_loss:
                status = TradeStatus.STOPPED_OUT.value
                exit_price = stop_loss
                exit_reason = ExitReason.STOP_LOSS.value
                pnl_rupees = (exit_price - entry_trigger) * shares
                pnl_pct = ((exit_price - entry_trigger) / entry_trigger) * 100.0

                ws.cell(row=row_idx, column=17, value=status)
                ws.cell(row=row_idx, column=19, value=round(pnl_pct, 2))
                ws.cell(row=row_idx, column=20, value=round(pnl_rupees, 2))
                ws.cell(row=row_idx, column=21, value=str(scan_date))
                ws.cell(row=row_idx, column=22, value=round(exit_price, 2))
                ws.cell(row=row_idx, column=23, value=exit_reason)
                updated_summary.append(f"🔴 *{symbol}*: Stop Loss Hit at ₹{stop_loss:.2f} (P&L: {pnl_pct:+.2f}%)")
                continue

            # 3. Check Target Exits
            if status in [TradeStatus.ACTIVE.value, TradeStatus.TARGET_1_HIT.value, TradeStatus.TARGET_2_HIT.value]:
                if high_p >= target_3:
                    status = TradeStatus.TARGET_3_HIT.value
                    exit_price = target_3
                    exit_reason = ExitReason.TARGET_3.value
                    pnl_rupees = (exit_price - entry_trigger) * shares
                    pnl_pct = ((exit_price - entry_trigger) / entry_trigger) * 100.0

                    ws.cell(row=row_idx, column=17, value=status)
                    ws.cell(row=row_idx, column=19, value=round(pnl_pct, 2))
                    ws.cell(row=row_idx, column=20, value=round(pnl_rupees, 2))
                    ws.cell(row=row_idx, column=21, value=str(scan_date))
                    ws.cell(row=row_idx, column=22, value=round(exit_price, 2))
                    ws.cell(row=row_idx, column=23, value=exit_reason)
                    updated_summary.append(f"🚀 *{symbol}*: Target 3 HIT at ₹{target_3:.2f} (P&L: {pnl_pct:+.2f}%)")
                    continue

                if high_p >= target_2 and status != TradeStatus.TARGET_2_HIT.value:
                    status = TradeStatus.TARGET_2_HIT.value
                    ws.cell(row=row_idx, column=17, value=status)
                    updated_summary.append(f"🎯 *{symbol}*: Target 2 HIT at ₹{target_2:.2f}")

                elif high_p >= target_1 and status == TradeStatus.ACTIVE.value:
                    status = TradeStatus.TARGET_1_HIT.value
                    ws.cell(row=row_idx, column=17, value=status)
                    updated_summary.append(f"🎯 *{symbol}*: Target 1 HIT at ₹{target_1:.2f}")

            # 4. Check 15-Session Time Stop
            if holding_sessions >= 15 and status not in [TradeStatus.TARGET_3_HIT.value, TradeStatus.STOPPED_OUT.value]:
                status = TradeStatus.TIME_EXPIRED.value
                exit_price = close_p
                exit_reason = ExitReason.TIME_STOP.value
                pnl_rupees = (exit_price - entry_trigger) * shares
                pnl_pct = ((exit_price - entry_trigger) / entry_trigger) * 100.0

                ws.cell(row=row_idx, column=17, value=status)
                ws.cell(row=row_idx, column=19, value=round(pnl_pct, 2))
                ws.cell(row=row_idx, column=20, value=round(pnl_rupees, 2))
                ws.cell(row=row_idx, column=21, value=str(scan_date))
                ws.cell(row=row_idx, column=22, value=round(exit_price, 2))
                ws.cell(row=row_idx, column=23, value=exit_reason)
                updated_summary.append(f"⏰ *{symbol}*: Time Stop (15 sessions) at ₹{close_p:.2f} (P&L: {pnl_pct:+.2f}%)")
                continue

            # Update unrealized P&L for active/target-hit trades
            unrealized_pnl_rupees = (close_p - entry_trigger) * shares
            unrealized_pnl_pct = ((close_p - entry_trigger) / entry_trigger) * 100.0
            ws.cell(row=row_idx, column=19, value=round(unrealized_pnl_pct, 2))
            ws.cell(row=row_idx, column=20, value=round(unrealized_pnl_rupees, 2))

        self._apply_sheet_formatting(ws)
        wb.save(self.file_path)
        self._sync_to_csv()

        summary_text = "\n".join(updated_summary) if updated_summary else "All open positions monitored cleanly. No level breaches today."
        logger.info(f"Excel ledger updated successfully for {scan_date}.")
        return summary_text

    def _apply_header_styles(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Formats the header row with dark blue fill and bold white text."""
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_num, col_name in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        ws.row_dimensions[1].height = 28

    def _apply_sheet_formatting(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Applies status color highlighting, number formats, borders, and auto-column widths."""
        self._apply_header_styles(ws)

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        status_fills = {
            TradeStatus.PENDING_ENTRY.value: PatternFill(start_color="FFF2CC", fill_type="solid"), # Light yellow
            TradeStatus.ACTIVE.value: PatternFill(start_color="DDEBF7", fill_type="solid"),        # Light blue
            TradeStatus.TARGET_1_HIT.value: PatternFill(start_color="E2EFDA", fill_type="solid"),  # Light green
            TradeStatus.TARGET_2_HIT.value: PatternFill(start_color="C6EFCE", fill_type="solid"),  # Green
            TradeStatus.TARGET_3_HIT.value: PatternFill(start_color="A9D08E", fill_type="solid"),  # Deep green
            TradeStatus.STOPPED_OUT.value: PatternFill(start_color="FCE4D6", fill_type="solid"),   # Light red
            TradeStatus.TIME_EXPIRED.value: PatternFill(start_color="F2F2F2", fill_type="solid"),  # Light grey
        }

        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
            status_val = str(ws.cell(row=row, column=17).value or "")

            for col in range(1, len(self.COLUMNS) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                # Highlight Status column
                if col == 17 and status_val in status_fills:
                    cell.fill = status_fills[status_val]
                    cell.font = Font(name="Calibri", size=10, bold=True)

                # Format P&L columns
                if col == 19 and isinstance(cell.value, (int, float)):
                    cell.number_format = "+0.00%;-0.00%;0.00%"
                    if cell.value > 0:
                        cell.font = Font(name="Calibri", size=10, color="375623", bold=True)
                    elif cell.value < 0:
                        cell.font = Font(name="Calibri", size=10, color="C00000", bold=True)

                # Format Currency columns
                if col in [8, 9, 10, 11, 12, 13, 15, 16, 20, 22] and isinstance(cell.value, (int, float)):
                    cell.number_format = "₹#,##0.00"

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    def _sync_to_csv(self) -> None:
        """Syncs Excel rows to CSV for lightweight parsing."""
        try:
            df = pd.read_excel(self.file_path, sheet_name="Trade Ledger")
            df.to_csv(self.csv_path, index=False)
        except Exception as e:
            logger.warning(f"Could not sync Excel ledger to CSV: {e}")
